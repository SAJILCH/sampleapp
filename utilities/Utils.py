import time

from appium.webdriver.common.mobileby import MobileBy

from dataProvider import configuration
import openpyxl
from appium.webdriver.common.touch_action import TouchAction
from hamcrest import none
from selenium.common.exceptions import NoSuchElementException, WebDriverException
from time import sleep

from selenium.webdriver.support.wait import WebDriverWait

from settings import *
from appium.webdriver import webelement


class PageUtils:

    def __init__(self, driver):
        self.driver = driver

    def get_element(self, locator):
        """
        Returns element based on provided locator.
        Locator include the method and locator value in a tuple.
        param locator:
        :return:
        """

        method = locator[0]
        values = locator[1]

        if type(values) is str:
            return self.get_element_by_type(method, values)
        elif type(values) is list:
            for value in values:
                try:
                    return self.get_element_by_type(method, value)
                except NoSuchElementException:
                    pass
            raise NoSuchElementException

    def get_element_by_type(self, method, value):
        if method == 'accessibility_id':
            return self.driver.find_element_by_accessibility_id(value)
        elif method == 'android':
            return self.driver.find_element_by_android_uiautomator('new UiSelector().%s' % value)
        elif method == 'ios':
            return self.driver.find_element_by_ios_uiautomation(value)
        elif method == 'class_name':
            return self.driver.find_element_by_class_name(value)
        elif method == 'id':
            return self.driver.find_element_by_id(value)
        elif method == 'xpath':
            return self.driver.find_element_by_xpath(value)
        elif method == 'name':
            return self.driver.find_element_by_name(value)
        else:
            raise Exception('Invalid locator method.')

    def get_elements(self, locator):
        """
        Returns element based on provided locator.
        Locator include the method and locator value in a tuple.
        param locator:
        :return:
        """

        method = locator[0]
        values = locator[1]

        if type(values) is str:
            return self.get_elements_by_type(method, values)
        elif type(values) is list:
            for value in values:
                try:
                    return self.get_elements_by_type(method, value)
                except NoSuchElementException:
                    pass
            raise NoSuchElementException

    def get_elements_by_type(self, method, value):
        if method == 'accessibility_id':
            return self.driver.find_elements_by_accessibility_id(value)
        elif method == 'android':
            return self.driver.find_elements_by_android_uiautomator(value)
        elif method == 'ios':
            return self.driver.find_elements_by_ios_uiautomation(value)
        elif method == 'class_name':
            return self.driver.find_elements_by_class_name(value)
        elif method == 'id':
            return self.driver.find_elements_by_id(value)
        elif method == 'xpath':
            return self.driver.find_elements_by_xpath(value)
        elif method == 'name':
            return self.driver.find_elements_by_name(value)
        else:
            raise Exception('Invalid locator method.')

        # element visible

    def is_visible(self, locator):
        time.sleep(2)
        try:
            el = self.get_element(locator).is_displayed()
            if el == True:
                return True
        except NoSuchElementException:
            return False

    def wait_visible(self, locator, timeout=2):
        i = 0
        while i != timeout:
            try:
                self.is_visible(locator)
                return self.get_element(locator)
            except NoSuchElementException:
                sleep(.5)
                i += 1
        raise Exception('Element never became visible: %s (%s)' % (locator[0], locator[1]))

        # clicks and taps

    def click(self, locator):
        element = self.wait_visible(locator)
        element.click()

    def tap(self, locator):
        element = self.wait_visible(locator)
        element.tap()

    def hide_keyboard(self):
        try:
            sleep(1)
            self.driver.hide_keyboard()
        except WebDriverException:
            pass

    def close(self):
        try:
            sleep(1)
            self.driver.back()
        except WebDriverException:
            pass

    def send_keys(self, locator, text):
        element = self.wait_visible(locator)
        element.send_keys(text)
        sleep(0.5)

    def clear(self, locator):
        element = self.wait_visible(locator)
        element.clear()
        sleep(0.5)

        # android scroll

    def android_scroll(self, locator):
        for _ in range(15):
            x = 1000
            try:
                if self.get_element(locator) != none:
                    value = True
                    if value is True:
                        break
            except NoSuchElementException:
                # swipe(start_x, start_y, end_x, end_y, duration)
                self.driver.swipe(500, 2000, 500, x, 500)
                self.driver.implicitly_wait(2)
                continue

    def ios_scroll(self, locator):
        el = self.wait_visible(locator)
        self.driver.execute_script('mobile: scroll', {"element": el, "toVisible": True})

        # common method to scroll in android & ios

    def scrolling(self, locator):
        if platform == 'ios':
            self.ios_scroll(locator)
        else:
            self.android_scroll(locator)
        # get text

    def get_text(self, locator):
        element = self.wait_visible(locator)
        return element.text

    def read_datas_from_excel(self, sheetname, cell_name):
        # dir_path = os.path.dirname(os.path.realpath(__file__))
        wb = openpyxl.load_workbook(filename=configuration.initial_setup.get("excel_path"))
        # wb = openpyxl.load_workbook("C:\\Users\\sharidas\\PycharmProjects\\Braintracker\\Data\\Data.xlsx")
        cell_value = wb[sheetname][cell_name].value
        return cell_value

    def read_excel(self):
        datalist = []
        wb = openpyxl.load_workbook(filename="C:\\Users\\sharidas\\PycharmProjects\\AskAnnie\\TestData\\data.xlsx")
        sh = wb['Demo']
        row_ct = sh.max_row
        col_ct = sh.max_column

        for i in range(2, row_ct + 1):
            row = []
            for j in range(1, col_ct + 1):
                row.append(sh.cell(row=i, column=j).value)
            datalist.append(row)
        return datalist

    def action_tap(self, locator):
        element = self.wait_visible(locator)
        actions = TouchAction(self.driver)
        actions.tap(element)
        actions.perform()

    def action_tap_bycordinates(self, x, y):
        # element = self.wait_visible(locator)
        actions = TouchAction(self.driver)
        actions.tap(x=x, y=y)
        actions.perform()

    def set_value(self, locator, text):
        element = self.wait_visible(locator)
        element.set_text(text)
        sleep(0.5)

    def is_element_present(self, locator):
        self.driver.implicitly_wait(15)

        if len(self.driver.find_elements(locator[0], locator[1])) != 0:
            print("element was present")
            return True
        else:
            print("element was not present")
            return False
        # ios scroll

    def ios_swipe(self, locator):
        el = self.wait_visible(locator)
        self.driver.execute_script('mobile: scroll', {"element": el, "toVisible": True})

        # common method to scroll in android & ios

    def swipe(self, locator):
        if platform == 'ios':
            self.ios_swipe(locator)
        else:
            self.android_swipe(locator)

    def get_system_bar(self):
        clock_visible = False
        try:
            for i in range(5):
                system_bars = self.driver.get_system_bars()
                print("System bars: ", system_bars)
                if system_bars['statusBar']['visible']:
                    clock_visible = True
                    break
                else:
                    time.sleep(1)
        except Exception as e:
            print("Error: ", e)

        if clock_visible:
            print("Clock is visible.")
            return True
        else:
            print("Clock is not visible.")
            return False

    def seekbarrow_scrolling(self, locator):
        if platform == 'ios':
            self.ios_scroll(locator)
        else:
            self.androidseekbarrow_scroll(locator)

    def androidseekbarrow_scroll(self, locator):
        for _ in range(15):
            x = 2700
            try:
                if self.get_element(locator) != none:
                    value = True
                    if value is True:
                        break
            except NoSuchElementException:
                # swipe(start_x, start_y, end_x, end_y, duration)
                self.driver.swipe(540, 2765, 540, x, 200)
                self.driver.implicitly_wait(2)
                continue

    def seekbarseat_scrolling(self, locator):
        if platform == 'ios':
            self.ios_scroll(locator)
        else:
            self.androidseekbarseat_scroll(locator)

    def androidseekbarseat_scroll(self, locator):
        for _ in range(15):
            x = 2700
            try:
                if self.get_element(locator) != none:
                    value = True
                    if value is True:
                        break
            except NoSuchElementException:
                # swipe(start_x, start_y, end_x, end_y, duration)
                self.driver.swipe(850, 2775, 850, x, 200)
                self.driver.implicitly_wait(2)
                continue

    def is_element_presenttest(self, locator):
        self.driver.implicitly_wait(15)
        print(locator)
        l = self.driver.find_elements(locator)
        print(l)
        full_xpath = l.get_attribute('content-desc')
        print(full_xpath)
        return full_xpath

    def change_accesbility(self):
        self.driver.start_activity('com.android.settings', '.Settings')
        accessibility_option = self.driver.find_element_by_android_uiautomator('new UiSelector().text("Accessibility")')
        accessibility_option.click()
        font_size_option = self.driver.find_element_by_android_uiautomator('new UiSelector().text("Font size")')
        font_size_option.click()
        largest_font_option = self.driver.find_element_by_android_uiautomator('new UiSelector().text("Largest")')
        largest_font_option.click()

