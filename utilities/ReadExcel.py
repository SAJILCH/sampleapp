import openpyxl
from dataProvider import configuration

def read_excel(sheet):
    datalist = []
    wb = openpyxl.load_workbook(filename=configuration.initial_setup.get("excel_path"))
    sh = wb[sheet]
    row_ct = sh.max_row
    col_ct = sh.max_column

    for i in range(2, row_ct + 1):
        row = []
        for j in range(1, col_ct + 1):
            row.append(sh.cell(row=i, column=j).value)
        datalist.append(row)
    return datalist


